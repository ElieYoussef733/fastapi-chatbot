import os
import json
import re
from typing import List, Dict

import numpy as np
from dotenv import load_dotenv
from openai import OpenAI

# PDF
import PyPDF2

# DOCX
from docx import Document as DocxDocument

# XLSX
from openpyxl import load_workbook

# Images + OCR
from PIL import Image
import pytesseract
import pytesseract

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

load_dotenv()

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
client = OpenAI(api_key=OPENAI_API_KEY)

UPLOADS_DIR = "knowledge/uploads"
INDEX_DIR = "knowledge/indexes"

# ------------------------------------------------
# TEXT EXTRACTION
# ------------------------------------------------

def extract_text_from_pdf(path: str) -> str:
    text_parts = []
    with open(path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            t = page.extract_text() or ""
            text_parts.append(t)
    return "\n".join(text_parts)


def extract_text_from_docx(path: str) -> str:
    doc = DocxDocument(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_text_from_txt(path: str) -> str:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        return f.read()


def extract_text_from_xlsx(path: str) -> str:
    wb = load_workbook(path, data_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            # join non-None cells as strings
            row_vals = [str(cell) for cell in row if cell is not None]
            if row_vals:
                texts.append(" | ".join(row_vals))
    return "\n".join(texts)


def extract_text_from_image(path: str) -> str:
    # Requires Tesseract installed on the system
    # and pytesseract configured
    image = Image.open(path)
    text = pytesseract.image_to_string(image, lang="eng")
    return text


def extract_text_from_file(path: str) -> str:
    lower = path.lower()
    if lower.endswith(".pdf"):
        return extract_text_from_pdf(path)
    elif lower.endswith(".docx"):
        return extract_text_from_docx(path)
    elif lower.endswith(".txt"):
        return extract_text_from_txt(path)
    elif lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return extract_text_from_xlsx(path)
    elif lower.endswith((".png", ".jpg", ".jpeg", ".bmp", ".tiff")):
        return extract_text_from_image(path)
    else:
        raise ValueError(f"Unsupported file type: {path}")


# ------------------------------------------------
# CLEAN + CHUNK
# ------------------------------------------------

def clean_text(text: str) -> str:
    # Remove null bytes
    text = text.replace("\x00", "")
    # Normalize whitespace
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def split_into_chunks(text: str, max_chars: int = 800, overlap: int = 100) -> List[str]:
    """
    Simple sliding-window chunking by characters.
    """
    text = text or ""
    text = clean_text(text)
    chunks = []
    start = 0
    length = len(text)

    if length == 0:
        return []

    while start < length:
        end = min(start + max_chars, length)
        chunk = text[start:end]
        chunks.append(chunk)
        if end == length:
            break
        start = end - overlap  # slide with overlap
        if start < 0:
            start = 0
    return chunks


# ------------------------------------------------
# EMBEDDINGS + INDEX
# ------------------------------------------------

def embed_texts(chunks: List[str]) -> List[Dict]:
    """
    Turn list of strings into list of {text, embedding} dicts.
    """
    if not chunks:
        return []

    # OpenAI embeddings has max batch size, but we can send all for small lists
    resp = client.embeddings.create(
        model="text-embedding-3-small",
        input=chunks
    )

    items = []
    for chunk, data in zip(chunks, resp.data):
        emb = data.embedding
        items.append({
            "text": chunk,
            "embedding": emb,
        })
    return items


def load_existing_index(business_id: str) -> List[Dict]:
    path = os.path.join(INDEX_DIR, f"{business_id}_kb.json")
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data


def save_index(business_id: str, items: List[Dict]):
    os.makedirs(INDEX_DIR, exist_ok=True)
    path = os.path.join(INDEX_DIR, f"{business_id}_kb.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False)


def merge_index(existing: List[Dict], new_items: List[Dict]) -> List[Dict]:
    """
    Simple append-based merge. You could add deduplication later.
    """
    return existing + new_items


# ------------------------------------------------
# PUBLIC FUNCTION: PROCESS A FILE FOR A BUSINESS
# ------------------------------------------------

def process_file_for_business(business_id: str, file_path: str) -> int:
    """
    Extract text, chunk, embed, and update KB for given business.
    Returns number of chunks added.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    # 1) Extract
    raw_text = extract_text_from_file(file_path)
    if not raw_text.strip():
        print(f"[WARN] No text extracted from {file_path}")
        return 0

    # 2) Chunk
    chunks = split_into_chunks(raw_text)
    if not chunks:
        print(f"[WARN] No chunks created from {file_path}")
        return 0

    # 3) Embed
    items = embed_texts(chunks)

    # 4) Load existing KB
    existing = load_existing_index(business_id)

    # 5) Merge
    merged = merge_index(existing, items)

    # 6) Save
    save_index(business_id, merged)

    return len(chunks)
